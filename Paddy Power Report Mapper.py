"""Paddy Power monthly report generator for Streamlit Cloud."""

from __future__ import annotations

import io
import re
import shutil
import subprocess
import tempfile
import zipfile
from collections import Counter, OrderedDict
from concurrent.futures import ThreadPoolExecutor
from copy import copy
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
import pandas as pd
import streamlit as st
from dateutil.relativedelta import relativedelta
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.formula import ArrayFormula


REQUIRED_LIVE_SHEETS = (
    "Checks",
    "Input",
    "Regions",
    "This Period",
    "YTD",
    "R12M",
    "Summary Data",
    "Store Performance",
    "Performance by Day & Time",
    "District Performance",
    "Operational Questions",
    "Performance Over Time",
)

NON_LIVE_SHEETS = (
    "Summary Data",
    "Store Performance",
    "Performance by Day & Time",
    "District Performance",
    "Operational Questions",
    "Performance Over Time",
    "Performance Over Time Chart",
)

VALUE_ONLY_SHEETS = (
    "Summary Data",
    "Store Performance",
    "Performance by Day & Time",
    "District Performance",
    "Operational Questions",
    "Performance Over Time",
)

DATE_OUTPUT_HEADERS = {
    "Order Deadline",
    "Order End Date",
    "Submitted Date",
    "Approved Date",
    "Actual Visit Date",
    "Date of Visit",
}

TIME_OUTPUT_HEADERS = {
    "Actual Visit Time",
    "ON ENTRY / BROWSING",
    "What was the time when you entered the shop?",
    "What time did you leave the shop?",
    "Please rate your overall customer service experience between 1-5 (where 1 is poor and 5 is excellent):",
}

EIRCODE_PATTERN = re.compile(r"^[A-Z]\d(?:\d|[A-Z])\s?[A-Z0-9]{4}$")
GB_POSTCODE_PATTERN = re.compile(r"^(?!BT)[A-Z]{1,2}\d{1,2}[A-Z]?\s?\d[A-Z]{2}$")
MONTH_IN_FILENAME = re.compile(
    r"Paddy Power (?:UK|ROI) Test Purchases - ([A-Za-z]+) (\d{4}) - LIVE\.xlsx$",
    re.IGNORECASE,
)


GB_MAPPING: "OrderedDict[str, Optional[str]]" = OrderedDict(
    [
        ("Order Number", "order_internal_id"),
        ("Client Name", "client_name"),
        ("Audit ID", "internal_id"),
        ("Site ID", "site_internal_id"),
        ("Order Deadline", "end_date"),
        ("Responsibility", "responsibility"),
        ("Premises Name", "site_name"),
        ("Address1", "site_address_1"),
        ("Address2", "site_address_2"),
        ("Address3", "site_address_3"),
        ("Post Code", "site_post_code"),
        ("Submitted Date", "submitted_date"),
        ("Approved Date", "approval_date"),
        ("Item To Order", "item_to_order"),
        ("Actual Visit Date", "date_of_visit"),
        ("Actual Visit Time", "time_of_visit"),
        ("AMPM", None),
        ("Pass-Fail", "primary_result"),
        ("Pass-Fail2", "secondary_result"),
        ("Abort Reason", "Please detail why you were unable to conduct this audit:"),
        ("Extra Site 1", "site_code"),
        ("Extra Site 2", None),
        ("Extra Site 3", None),
        ("Extra Site 4", "Were you challenged for ID on entry, at the machine, after machine play, or at the counter?"),
        ("Till ID?", None),
        ("VISITORSEX", None),
        ("ON ENTRY / BROWSING", "What was the time when you entered the shop?"),
        ("As you entered the shop was eye contact made by a member of staff?", "As you entered the shop was eye contact made by a member of staff?"),
        ("Were you acknowledged by any staff members?", "Were you acknowledged by any staff members?"),
        ("Were you challenged for ID on entry, at the machine or after machine play? (Please indicate below at which point of your visit you were challenged):", "Were you challenged for ID on entry, at the machine, after machine play, or at the counter?"),
        ("Were you asked for ID before or after you put a coin into the machine?", "Were you asked for ID before or after you put a coin into the machine?"),
        ("Please accurately describe the staff member who asked you for ID at one of these points:", "Please accurately describe the staff member who asked you for ID at one of these points:"),
        ("Was the staff member wearing a name badge?", "Was the staff member wearing a name badge?"),
        ("Was the member of staff wearing a (black) Paddy Power uniform?", "What was the name of the staff member?"),
        ("Please describe the manner in which you were challenged and add any other comments you feel are relevant:", "Please describe the manner in which you were challenged and add any other comments you feel are relevant:"),
        ("MACHINE AREA", None),
        ("Did all the gaming machines appear to be working?", "Did all the gaming machines appear to be working?"),
        ("Were all the machines visible from the counter?", "Were all the machines visible from the counter?"),
        ("Please describe what the staff member was doing as you approached the counter:", "Please describe what the staff member was doing as you approached the counter:"),
        ("Did the staff member who served you make eye contact with you during the transaction?", "Did the staff member who served you make eye contact with you during the transaction?"),
        ("When was eye contact first made?", "When was eye contact first made?"),
        ('Were "Think 21" signs visible in the machine area?', None),
        ("PLACING THE BET", None),
        ("Please describe what the server was doing as you approached the counter (e.g. serving a customer, talking to colleagues):", None),
        ("At the till, did the person who served you ask your age?", None),
        ("Did the staff member who served you at the till ask for ID?", None),
        ("Please enter the 17 digit number from your betting slip:", "Please enter the 17 digit number from your betting slip:"),
        ("Unnamed: 47", None),
        ("How many staff were visible in the shop at the time of your visit?", "How many staff were visible in the shop at the time of your visit?"),
        ("When were the staff first aware of you in the shop?", "When were the staff first aware of you in the shop?"),
        ("How many customers were in the shop at the time of your audit?", "How many customers were in the shop at the time of your audit?"),
        ("Did you see any 'Think 21' posters in the shop?", "Did you see any 'Think 25' posters in the shop?"),
        ("Did you see any 'Think 21' posters behind the counter?", "Did you see any 'Think 25' posters behind the counter?"),
        ("Please give a detailed report of your audit, providing a full description of your experience from entering to leaving the shop:", "Please give a detailed report of your audit, providing a full description of your experience from entering to leaving the shop:"),
        ("What time did you leave the shop?", "What time did you leave the shop?"),
        ("Were you wearing a face mask/covering during the audit?", None),
        ("Please confirm in the space below whether or not you were asked for ID:", None),
        ("Unnamed: 57", None),
        ("Unnamed: 58", "Please confirm below whether or not you were asked for ID:"),
        ("Unnamed: 59", None),
        ("Unnamed: 60", None),
        ("Unnamed: 61", None),
        ("Unnamed: 62", "Please confirm below whether or not you were asked for ID:"),
    ]
)


IE_MAPPING: "OrderedDict[str, Optional[str]]" = OrderedDict(
    [
        ("Order Number", "order_internal_id"),
        ("Client Name", "client_name"),
        ("Audit ID", "internal_id"),
        ("Site ID", "site_internal_id"),
        ("Order End Date", "end_date"),
        ("Responsibility", "responsibility"),
        ("Site Name", "site_name"),
        ("Address 1", "site_address_1"),
        ("Address 2", "site_address_2"),
        ("Address 3", "site_address_3"),
        ("Post Code", "site_post_code"),
        ("Submitted Date", "submitted_date"),
        ("Approved Date", "approval_date"),
        ("Item To Order", "item_to_order"),
        ("Date of Visit", "date_of_visit"),
        ("Actual Visit Time", "time_of_visit"),
        ("AMPM", None),
        ("Pass-Fail", "primary_result"),
        ("Pass-Fail2", "secondary_result"),
        ("Abort Reason", "Please detail why you were unable to conduct this audit:"),
        ("Extra Site 1", "site_code"),
        ("Unnamed: 21", None),
        ("Unnamed: 22", None),
        ("Were you challenged for ID on entry, at the machine, after machine play, or at the counter?", "Were you challenged for ID on entry, at the machine, after machine play, or at the counter?"),
        ("Unnamed: 24", None),
        ("Unnamed: 25", None),
        ("Unnamed: 26", None),
        ("What was the time when you entered the shop?", "What was the time when you entered the shop?"),
        ("As you entered the shop was eye contact made by a member of staff?", "As you entered the shop was eye contact made by a member of staff?"),
        ("Were you acknowledged by any staff members?", "Were you acknowledged by any staff members?"),
        ("Please describe any acknowledgement by staff members:", "Please describe any acknowledgement by staff members:"),
        ("Please explain what may have prevented staff from greeting you:", "Please explain what may have prevented staff from greeting you:"),
        ("If so, what was their name?", "Were you challenged for ID on entry, at the machine, after machine play, or at the counter?"),
        ("Was the member of staff wearing a (black) Paddy Power uniform?", "Were you asked for ID before or after you put a coin into the machine?"),
        ("Please describe the manner in which you were challenged and add any other comments you feel are relevant:", "Please describe the manner in which you were challenged and add any other comments you feel are relevant:"),
        ("Did the staff member appear to record any of the details from your ID?", "Did the staff member appear to record any of the details from your ID?"),
        ("Please accurately describe the staff member who asked you for ID at one of these points:", "Please accurately describe the staff member who asked you for ID at one of these points:"),
        ("If not, please state why:", "Was the staff member who served you wearing a name badge?"),
        ("Did all the gaming machines appear to be working?", "As required, did you browse for 2 minutes, including time at the self-service machine?"),
        ("Were all the machines visible from the counter?", "Please explain why you did not browse for 2 minutes:"),
        ("Please describe what the staff member was doing as you approached the counter:", "Please describe what the staff member was doing as you approached the counter:"),
        ("Did the staff member who served you make eye contact with you?", "Did the staff member who served you make eye contact with you?"),
        ("When was eye contact first made?", "When was eye contact first made?"),
        ("Please accurately describe the staff member who served you:", "Please accurately describe the staff member who served you:"),
        ("Did the staff member who served you smile?", "Did the staff member who served you smile?"),
        ("Did the staff member who served you greet you?", "Did the staff member who served you greet you?"),
        ("Was the staff member who served you wearing a name badge?", "Was the staff member who served you wearing a name badge?"),
        ("What was the name of the staff member who served you?", "What was the name of the staff member who served you?"),
        ("Please enter the 17 digit number from the bottom of your betting slip:", "Please enter the 17 digit number from the bottom of your betting slip:"),
        ("How many staff were on duty in the shop at the time of your audit?", "How many staff were on duty in the shop at the time of your audit?"),
        ("Was the staff member wearing a shirt and tie or a shirt and cravat, as shown in the briefing document?", "Was the staff member wearing a shirt and tie or a shirt and cravat, as shown in the briefing document?"),
        ("Describe what the staff member was wearing:", "Describe what the staff member was wearing:"),
        ("When were the staff first aware of you in the shop?", "When were the staff first aware of you in the shop?"),
        ("How many customers were in the shop at the time of your audit?", "How many customers were in the shop at the time of your audit?"),
        ("Did you see any 'Think 21' signage on the front door of the shop?", "Did you see any 'Think 21' signage on the front door of the shop?"),
        ("Did you see any 'Think 21' posters in the shop?", "Did you see any 'Think 21' posters in the shop?"),
        ("Did you see any 'Think 21' behind the counter?", "Did you see any 'Think 21' behind the counter?"),
        ("Please give a detailed report of your audit, providing a full description of your experience from entering to leaving the shop:", "Please give a detailed report of your audit, providing a full description of your experience from entering to leaving the shop:"),
        ("Unnamed: 58", "Please rate your overall customer service experience between 1-5 (where 1 is poor and 5 is excellent):"),
        ("Please rate your overall customer service experience between 1-5 (where 1 is poor and 5 is excellent):", "What time did you leave the shop?"),
        ("Were you wearing a face mask/covering during the audit?", None),
        ("Were you asked to remove your mask/covering during the audit?", None),
        ("Please use this space to explain anything unusual about your visit or to clarify any detail of your report:", "Please use this space to explain anything unusual about your visit or to clarify any detail of your report:"),
        ("As required, did you browse for 2 minutes, including time at the self-service machine?", "As required, did you browse for 2 minutes, including time at the self-service machine?"),
        ("Please explain why you did not browse for 2 minutes:", "Please explain why you did not browse for 2 minutes:"),
        ("Please confirm below whether or not you were asked for ID:", "Please confirm below whether or not you were asked for ID:"),
    ]
)


STORE_DB_SHEETS = {
    "UK": ("GB Shop List",),
    "ROI": ("NI Shop List", "Ire Shop List"),
}


def normalise_header(value: object) -> str:
    return re.sub(r"\s+", " ", "" if value is None else str(value).strip()).casefold()


def normalise_shop_code(value: object) -> object:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)) and float(value).is_integer():
        return int(value)
    text = str(value).strip()
    if re.fullmatch(r"\d+", text):
        return int(text)
    return text


def shop_code_key(value: object) -> str:
    normalised = normalise_shop_code(value)
    return "" if normalised is None else str(normalised).strip()


def store_records_from_sheet(workbook, sheet_name: str) -> List[Dict[str, object]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"The Store DB is missing the required '{sheet_name}' sheet.")
    sheet = workbook[sheet_name]
    headers = {
        normalise_header(cell.value): cell.column
        for cell in sheet[1]
        if normalise_header(cell.value)
    }
    required = ("shop no.", "site internal id", "district", "shop name", "district manager")
    missing = [header for header in required if header not in headers]
    if missing:
        raise ValueError(
            f"The '{sheet_name}' sheet in the Store DB is missing column(s): {', '.join(missing)}."
        )

    records: List[Dict[str, object]] = []
    for row in range(2, sheet.max_row + 1):
        shop_no = normalise_shop_code(sheet.cell(row, headers["shop no."]).value)
        if shop_no in (None, ""):
            continue
        district = str(sheet.cell(row, headers["district"]).value or "").strip()
        shop_name = str(sheet.cell(row, headers["shop name"]).value or "").strip()
        if not district or not shop_name:
            raise ValueError(
                f"The '{sheet_name}' sheet has a blank District or Shop Name for shop {shop_no}."
            )
        records.append(
            {
                "shop_no": shop_no,
                "site_internal_id": str(
                    sheet.cell(row, headers["site internal id"]).value or ""
                ).strip(),
                "district": district,
                "shop_name": shop_name,
                "district_manager": str(
                    sheet.cell(row, headers["district manager"]).value or ""
                ).strip(),
            }
        )
    return records


def load_store_database(data: bytes) -> Dict[str, List[Dict[str, object]]]:
    try:
        workbook = load_workbook_from_bytes(data, data_only=True)
    except Exception as exc:
        raise ValueError(f"The Paddy Power Store DB could not be opened: {exc}") from exc

    result: Dict[str, List[Dict[str, object]]] = {}
    for country, sheets in STORE_DB_SHEETS.items():
        records: List[Dict[str, object]] = []
        for sheet_name in sheets:
            records.extend(store_records_from_sheet(workbook, sheet_name))
        keys = [shop_code_key(record["shop_no"]) for record in records]
        duplicates = sorted(key for key, count in Counter(keys).items() if count > 1)
        if duplicates:
            raise ValueError(
                f"The Store DB contains duplicate {country} shop code(s): {', '.join(duplicates)}."
            )
        result[country] = records
    return result


def classify_country(postcode: object) -> str:
    pc = "" if postcode is None else str(postcode).strip().upper()
    pc = re.sub(r"\s+", " ", pc)
    if not pc or pc.startswith("BT") or EIRCODE_PATTERN.fullmatch(pc):
        return "IE"
    if GB_POSTCODE_PATTERN.fullmatch(pc):
        return "GB"
    return "IE"


def parse_date_value(value: object) -> Optional[datetime]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time())
    text = str(value).strip()
    if not text or text.lower() == "invalid date":
        return None
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime().replace(tzinfo=None)


def parse_time_value(value: object) -> Optional[time]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None)
    if isinstance(value, time):
        return value.replace(tzinfo=None)
    text = str(value).strip()
    if not text or text.lower() == "invalid date":
        return None
    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime().time().replace(tzinfo=None)


def coerce_general_value(value: object) -> object:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    if not text or text.lower() == "invalid date":
        return None
    if text == "True":
        return True
    if text == "False":
        return False
    integer_match = re.fullmatch(r"-?(?:0|[1-9]\d*)", text)
    if (
        integer_match
        and not (len(text) > 1 and text.startswith("0"))
        and len(text.lstrip("-")) <= 15
    ):
        try:
            return int(text)
        except ValueError:
            pass
    if re.fullmatch(r"-?(?:0|[1-9]\d*)\.\d+", text):
        try:
            return float(text)
        except ValueError:
            pass
    return text


def mapped_rows(df: pd.DataFrame, mapping: "OrderedDict[str, Optional[str]]") -> List[List[object]]:
    rows: List[List[object]] = []
    for record in df.to_dict(orient="records"):
        row: List[object] = []
        for output_header, export_header in mapping.items():
            raw = record.get(export_header, "") if export_header else ""
            if output_header in DATE_OUTPUT_HEADERS:
                row.append(parse_date_value(raw))
            elif output_header in TIME_OUTPUT_HEADERS:
                row.append(parse_time_value(raw))
            else:
                row.append(coerce_general_value(raw))
        rows.append(row)
    return rows


def load_workbook_from_bytes(data: bytes, *, data_only: bool = False):
    return openpyxl.load_workbook(
        io.BytesIO(data),
        data_only=data_only,
        read_only=False,
        keep_links=True,
    )


def remove_invalid_defined_names(data: bytes) -> bytes:
    """Remove defined names that Excel would repair after source sheets are removed.

    The non-LIVE workbook intentionally contains only presentation sheets. Names
    that still point at deleted calculation/input sheets, external workbook names,
    or #REF! targets are invalid in that reduced workbook. Valid print titles,
    filters, and names on retained sheets are preserved.
    """
    from xml.etree import ElementTree as ET

    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    sheet_ref_pattern = re.compile(r"(?:(?:'((?:[^']|'')+)')|([^'!,\s]+))!")
    source = io.BytesIO(data)
    output = io.BytesIO()
    with zipfile.ZipFile(source, "r") as input_archive:
        root = ET.fromstring(input_archive.read("xl/workbook.xml"))
        valid_sheets = {
            node.attrib.get("name", "")
            for node in root.findall(f".//{{{main_ns}}}sheet")
        }
        for container in list(root.findall(f"{{{main_ns}}}definedNames")):
            for defined_name in list(container.findall(f"{{{main_ns}}}definedName")):
                target = defined_name.text or ""
                referenced_sheets = {
                    (quoted or plain or "").replace("''", "'")
                    for quoted, plain in sheet_ref_pattern.findall(target)
                }
                invalid = (
                    "#REF!" in target.upper()
                    or "[" in target
                    or bool(referenced_sheets - valid_sheets)
                )
                local_sheet_id = defined_name.attrib.get("localSheetId")
                if local_sheet_id is not None:
                    try:
                        invalid = invalid or int(local_sheet_id) >= len(valid_sheets)
                    except ValueError:
                        invalid = True
                if invalid:
                    container.remove(defined_name)
            if not list(container):
                root.remove(container)
        workbook_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        with zipfile.ZipFile(output, "w") as output_archive:
            for info in input_archive.infolist():
                content = (
                    workbook_xml
                    if info.filename == "xl/workbook.xml"
                    else input_archive.read(info.filename)
                )
                output_archive.writestr(info, content)
    return output.getvalue()


def workbook_to_bytes(
    workbook,
    drawing_source: Optional[bytes] = None,
    external_link_source: Optional[bytes] = None,
) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    data = buffer.getvalue()
    if drawing_source:
        data = restore_package_parts(data, drawing_source, ("xl/drawings/", "xl/media/"))
    if external_link_source:
        # openpyxl drops the alternate URL relationship used by externalLink2.xml,
        # which makes Excel repair the workbook on open. These legacy dashboard
        # links are unchanged, so retain their original, internally consistent XML.
        data = restore_package_parts(data, external_link_source, ("xl/externalLinks/",))
    return normalise_drawing_relationships(data)


def restore_package_parts(data: bytes, source_data: bytes, prefixes: Sequence[str]) -> bytes:
    """Restore package parts that must remain byte-for-byte compatible with the template."""
    output = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(data), "r") as input_archive,
        zipfile.ZipFile(io.BytesIO(source_data), "r") as source_archive,
        zipfile.ZipFile(output, "w") as output_archive,
    ):
        for info in input_archive.infolist():
            if not info.filename.startswith(prefixes):
                output_archive.writestr(info, input_archive.read(info.filename))
        for info in source_archive.infolist():
            if info.filename.startswith(prefixes):
                output_archive.writestr(info, source_archive.read(info.filename))
    return output.getvalue()


def region_records_from_workbook(workbook) -> List[Dict[str, object]]:
    sheet = workbook["Regions"]
    records: List[Dict[str, object]] = []
    for row in range(2, sheet.max_row + 1):
        shop_no = normalise_shop_code(sheet.cell(row, 1).value)
        district = str(sheet.cell(row, 2).value or "").strip()
        shop_name = str(sheet.cell(row, 3).value or "").strip()
        if shop_no in (None, "") or not district or not shop_name:
            continue
        manager = str(sheet.cell(row, 4).value or "").strip()
        if manager == district:
            manager = ""
        records.append(
            {
                "shop_no": shop_no,
                "site_internal_id": "",
                "district": district,
                "shop_name": shop_name,
                "district_manager": manager,
            }
        )
    return records


def active_shop_codes(records: Sequence[Sequence[object]], mapping) -> set[str]:
    site_code_index = list(mapping.keys()).index("Extra Site 1")
    return {
        shop_code_key(record[site_code_index])
        for record in records
        if shop_code_key(record[site_code_index])
    }


def merged_store_records(
    workbook,
    store_db_records: Sequence[Dict[str, object]],
    retained_records: Sequence[Sequence[object]],
    required_records: Sequence[Sequence[object]],
    mapping,
) -> List[Dict[str, object]]:
    """Use the Store DB for current shops and retain historical shops still in R12M/YTD."""
    current_keys = {shop_code_key(record["shop_no"]) for record in store_db_records}
    needed_legacy_keys = active_shop_codes(retained_records, mapping) - current_keys
    old_records = region_records_from_workbook(workbook)
    old_by_key = {shop_code_key(record["shop_no"]): record for record in old_records}
    missing = sorted(
        active_shop_codes(required_records, mapping) - current_keys - set(old_by_key)
    )
    if missing:
        raise ValueError(
            "The Store DB and previous LIVE report do not contain shop code(s) used by the report data: "
            + ", ".join(missing)
            + "."
        )
    legacy = [
        record
        for record in old_records
        if shop_code_key(record["shop_no"]) in needed_legacy_keys
    ]
    return [dict(record) for record in store_db_records] + legacy


def update_regions_sheet(workbook, records: Sequence[Dict[str, object]]) -> int:
    sheet = workbook["Regions"]
    last_existing = max(
        (row for row in range(2, sheet.max_row + 1) if sheet.cell(row, 1).value not in (None, "")),
        default=1,
    )
    last_target = 1 + len(records)
    clear_last = max(last_existing, last_target)
    for row in range(1, clear_last + 1):
        for column in range(1, 6):
            sheet.cell(row, column).value = None

    headers = ("Shop No.", "District", "Shop Name", "District Manager")
    for column, header in enumerate(headers, start=1):
        sheet.cell(1, column).value = header
    for target_row, record in enumerate(records, start=2):
        if target_row > last_existing:
            copy_row_style(sheet, 2, target_row, 5)
        sheet.cell(target_row, 1).value = record["shop_no"]
        sheet.cell(target_row, 2).value = record["district"]
        sheet.cell(target_row, 3).value = record["shop_name"]
        sheet.cell(target_row, 4).value = record["district_manager"]

    sheet.auto_filter.ref = f"A1:D{last_target}"
    return last_target


def update_region_lookup_ranges(workbook, last_region_row: int) -> None:
    pattern = re.compile(r"Regions!\$A\$1:\$B\$\d+", re.IGNORECASE)
    replacement = f"Regions!$A$1:$B${last_region_row}"
    for sheet_name in ("This Period", "YTD", "R12M"):
        sheet = workbook[sheet_name]
        for cell in sheet[4]:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = pattern.sub(replacement, cell.value)


def make_abort_result_formulas_explicit(workbook) -> None:
    """Ensure abort rows return ABORT instead of looking up an empty challenge key."""
    for sheet_name in ("This Period", "YTD", "R12M"):
        sheet = workbook[sheet_name]
        for cell in sheet[4]:
            header = sheet.cell(3, cell.column).value
            formula = cell.value
            if (
                header in ("Challenge", "Basic", "PP Result")
                and isinstance(formula, str)
                and formula.startswith("=VLOOKUP(")
                and "Input!$Q$" in formula
            ):
                cell.value = f'=IF(UPPER($S4)="ABORT","ABORT",{formula[1:]})'


def snapshot_row(sheet, row: int, max_column: int) -> Dict[str, Any]:
    return {
        "source_row": row,
        "values": [sheet.cell(row, column).value for column in range(1, max_column + 1)],
        "styles": [copy(sheet.cell(row, column)._style) for column in range(1, max_column + 1)],
        "height": sheet.row_dimensions[row].height,
        "hidden": sheet.row_dimensions[row].hidden,
    }


def restore_snapshot_row(
    sheet,
    target_row: int,
    snapshot: Dict[str, Any],
    *,
    translate_formulas: bool = False,
) -> None:
    sheet.row_dimensions[target_row].height = snapshot["height"]
    sheet.row_dimensions[target_row].hidden = snapshot["hidden"]
    source_row = snapshot["source_row"]
    for column, (value, style) in enumerate(zip(snapshot["values"], snapshot["styles"]), start=1):
        cell = sheet.cell(target_row, column)
        cell._style = copy(style)
        if translate_formulas and isinstance(value, str) and value.startswith("="):
            source_coordinate = sheet.cell(source_row, column).coordinate
            cell.value = Translator(value, origin=source_coordinate).translate_formula(cell.coordinate)
        else:
            cell.value = value


def set_total_formulas(sheet, row: int, start_row: int, end_row: int) -> None:
    sheet.cell(row, 1).value = "Total"
    for column in (4, 5, 6, 7, 9, 10, 11, 12, 14, 15, 16, 17):
        letter = sheet.cell(row, column).column_letter
        sheet.cell(row, column).value = f"=SUM({letter}{start_row}:{letter}{end_row})"
    sheet.cell(row, 8).value = f'=IF(E{row}=0,"-",G{row}/E{row})'
    sheet.cell(row, 13).value = f'=IF(J{row}=0,"-",L{row}/J{row})'
    sheet.cell(row, 18).value = f'=IF(O{row}=0,"-",Q{row}/O{row})'


def rebuild_store_performance(workbook, records: Sequence[Dict[str, object]]) -> None:
    sheet = workbook["Store Performance"]
    start_row = 7
    old_total_row = next(
        (row for row in range(start_row, sheet.max_row + 1) if sheet.cell(row, 1).value == "Total"),
        sheet.max_row,
    )
    data_snapshot = snapshot_row(sheet, start_row, 18)
    total_snapshot = snapshot_row(sheet, old_total_row, 18)
    last_data_row = start_row + len(records) - 1
    new_total_row = last_data_row + 2
    clear_last = max(old_total_row, new_total_row)
    for row in range(start_row, clear_last + 1):
        for column in range(1, 19):
            sheet.cell(row, column).value = None

    for row, record in enumerate(records, start=start_row):
        restore_snapshot_row(sheet, row, data_snapshot, translate_formulas=True)
        sheet.cell(row, 1).value = record["shop_no"]
        sheet.cell(row, 2).value = record["shop_name"]

    restore_snapshot_row(sheet, new_total_row, total_snapshot)
    set_total_formulas(sheet, new_total_row, start_row, last_data_row)
    sheet.auto_filter.ref = f"A6:R{last_data_row}"


def find_row_with_value(sheet, value: object, start_row: int = 1) -> int:
    for row in range(start_row, sheet.max_row + 1):
        if sheet.cell(row, 1).value == value:
            return row
    raise ValueError(f"The '{sheet.title}' sheet is missing its '{value}' row.")


def ordered_districts(
    sheet,
    first_data_row: int,
    first_total_row: int,
    records: Sequence[Dict[str, object]],
) -> List[str]:
    available = {str(record["district"]).strip() for record in records}
    result: List[str] = []
    for row in range(first_data_row, first_total_row):
        label = str(sheet.cell(row, 1).value or "").strip()
        if label == "NI/ROI Mid North":
            for district in ("Northern Ireland", "ROI Mid North"):
                if district in available and district not in result:
                    result.append(district)
        elif label in available and label not in result:
            result.append(label)
    for record in records:
        district = str(record["district"]).strip()
        if district and district not in result:
            result.append(district)
    return result


def set_district_total_formulas(sheet, row: int, start_row: int, end_row: int) -> None:
    sheet.cell(row, 1).value = "Total"
    for column in (2, 3, 4, 5, 7, 8, 9, 10, 12, 13, 14, 15):
        letter = sheet.cell(row, column).column_letter
        sheet.cell(row, column).value = f"=SUM({letter}{start_row}:{letter}{end_row})"
    sheet.cell(row, 6).value = f'=IF(C{row}=0,"-",E{row}/C{row})'
    sheet.cell(row, 11).value = f'=IF(H{row}=0,"-",J{row}/H{row})'
    sheet.cell(row, 16).value = f'=IF(M{row}=0,"-",O{row}/M{row})'


def rebuild_district_performance(workbook, records: Sequence[Dict[str, object]]) -> List[str]:
    sheet = workbook["District Performance"]
    overall_title_row = find_row_with_value(sheet, "Overall Pass Rate")
    first_header_row = find_row_with_value(sheet, "District", overall_title_row)
    first_data_row = first_header_row + 1
    first_total_row = find_row_with_value(sheet, "Total", first_data_row)
    second_title_row = find_row_with_value(sheet, "Paddy Power Pass Rate", first_total_row + 1)
    second_header_row = find_row_with_value(sheet, "District", second_title_row)
    second_data_row = second_header_row + 1
    second_total_row = find_row_with_value(sheet, "Total", second_data_row)

    districts = ordered_districts(sheet, first_data_row, first_total_row, records)
    first_data_snapshot = snapshot_row(sheet, first_data_row, 16)
    first_total_snapshot = snapshot_row(sheet, first_total_row, 16)
    second_title_snapshot = snapshot_row(sheet, second_title_row, 16)
    second_period_snapshot = snapshot_row(sheet, second_title_row + 1, 16)
    second_header_snapshot = snapshot_row(sheet, second_header_row, 16)
    second_data_snapshot = snapshot_row(sheet, second_data_row, 16)
    second_total_snapshot = snapshot_row(sheet, second_total_row, 16)

    new_first_end = first_data_row + len(districts) - 1
    new_first_total = new_first_end + 2
    new_second_title = new_first_total + 3
    new_second_header = new_second_title + 2
    new_second_data = new_second_header + 1
    new_second_end = new_second_data + len(districts) - 1
    new_second_total = new_second_end + 2

    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row >= second_title_row:
            sheet.unmerge_cells(str(merged_range))
    clear_last = max(second_total_row, new_second_total)
    for row in range(first_data_row, clear_last + 1):
        for column in range(1, 17):
            sheet.cell(row, column).value = None

    for row, district in enumerate(districts, start=first_data_row):
        restore_snapshot_row(sheet, row, first_data_snapshot, translate_formulas=True)
        sheet.cell(row, 1).value = district
    restore_snapshot_row(sheet, new_first_total, first_total_snapshot)
    set_district_total_formulas(sheet, new_first_total, first_data_row, new_first_end)

    restore_snapshot_row(sheet, new_second_title, second_title_snapshot)
    restore_snapshot_row(sheet, new_second_title + 1, second_period_snapshot)
    restore_snapshot_row(sheet, new_second_header, second_header_snapshot)
    for start_column, end_column in ((2, 6), (7, 11), (12, 16)):
        sheet.merge_cells(
            start_row=new_second_title + 1,
            start_column=start_column,
            end_row=new_second_title + 1,
            end_column=end_column,
        )
    for row, district in enumerate(districts, start=new_second_data):
        restore_snapshot_row(sheet, row, second_data_snapshot, translate_formulas=True)
        sheet.cell(row, 1).value = district
    restore_snapshot_row(sheet, new_second_total, second_total_snapshot)
    set_district_total_formulas(sheet, new_second_total, new_second_data, new_second_end)
    return districts


def normalise_drawing_relationships(data: bytes) -> bytes:
    """Use portable relative drawing targets, matching the supplied Excel templates."""
    source = io.BytesIO(data)
    output = io.BytesIO()
    with zipfile.ZipFile(source, "r") as input_archive, zipfile.ZipFile(output, "w") as output_archive:
        for info in input_archive.infolist():
            content = input_archive.read(info.filename)
            if info.filename.startswith("xl/drawings/_rels/") and info.filename.endswith(".rels"):
                content = content.replace(b'Target="/xl/media/', b'Target="../media/')
                content = content.replace(b'Target="/xl/charts/', b'Target="../charts/')
            output_archive.writestr(info, content)
    return output.getvalue()


def validate_live_workbook(workbook, expected_country: str) -> None:
    missing = [name for name in REQUIRED_LIVE_SHEETS if name not in workbook.sheetnames]
    if missing:
        raise ValueError(f"The {expected_country} LIVE report is missing sheet(s): {', '.join(missing)}.")
    expected_g = "Premises Name" if expected_country == "UK" else "Site Name"
    actual_g = workbook["This Period"]["G3"].value
    if actual_g != expected_g:
        other = "ROI" if expected_country == "UK" else "UK"
        raise ValueError(
            f"The file uploaded as the {expected_country} LIVE report appears to be the {other} workbook."
        )


def parse_month_from_name(filename: str) -> Optional[date]:
    match = MONTH_IN_FILENAME.search(filename)
    if not match:
        return None
    try:
        return datetime.strptime(f"{match.group(1)} {match.group(2)}", "%B %Y").date().replace(day=1)
    except ValueError:
        return None


def report_month_from_workbook(workbook, filename: str) -> Optional[date]:
    try:
        value = workbook["Checks"]["B20"].value
        parsed = parse_date_value(value)
        if parsed:
            return parsed.date().replace(day=1)
    except Exception:
        pass
    return parse_month_from_name(filename)


def cached_report_month(data: bytes, filename: str) -> Optional[date]:
    try:
        workbook = load_workbook_from_bytes(data, data_only=True)
        value = workbook["Checks"]["B20"].value
        parsed = parse_date_value(value)
        if parsed:
            return parsed.date().replace(day=1)
    except Exception:
        pass
    return parse_month_from_name(filename)


def export_report_month(df: pd.DataFrame) -> date:
    if "date_of_visit" not in df.columns:
        raise ValueError("The data export is missing the required 'date_of_visit' column.")
    parsed = pd.to_datetime(df["date_of_visit"], dayfirst=True, errors="coerce")
    months = sorted({(d.year, d.month) for d in parsed.dropna()})
    if not months:
        raise ValueError("No valid visit dates were found in 'date_of_visit'.")
    if len(months) != 1:
        formatted = ", ".join(datetime(y, m, 1).strftime("%B %Y") for y, m in months)
        raise ValueError(f"The export contains more than one visit month: {formatted}.")
    return date(months[0][0], months[0][1], 1)


def audit_ids(workbook, sheet_name: str) -> set[str]:
    sheet = workbook[sheet_name]
    ids: set[str] = set()
    for row in range(4, sheet.max_row + 1):
        value = sheet.cell(row, 3).value
        if value not in (None, ""):
            ids.add(str(value).strip())
    return ids


def raw_records(workbook, sheet_name: str, raw_width: int) -> List[List[object]]:
    sheet = workbook[sheet_name]
    records: List[List[object]] = []
    for row in range(4, sheet.max_row + 1):
        if sheet.cell(row, 3).value in (None, ""):
            continue
        records.append([sheet.cell(row, column).value for column in range(1, raw_width + 1)])
    return records


def record_visit_date(record: Sequence[object]) -> Optional[date]:
    parsed = parse_date_value(record[14] if len(record) > 14 else None)
    return parsed.date() if parsed else None


def formula_columns(sheet) -> Dict[int, str]:
    result: Dict[int, str] = {}
    for cell in sheet[4]:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            result[cell.column] = cell.value
    return result


def copy_row_style(sheet, source_row: int, target_row: int, max_column: int) -> None:
    if source_row == target_row:
        return
    if source_row in sheet.row_dimensions:
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
        sheet.row_dimensions[target_row].hidden = sheet.row_dimensions[source_row].hidden
    for column in range(1, max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format


def write_data_sheet(
    sheet,
    records: Sequence[Sequence[object]],
    raw_width: int,
) -> None:
    formulas = formula_columns(sheet)
    existing_last = max(
        (row for row in range(4, sheet.max_row + 1) if sheet.cell(row, 3).value not in (None, "")),
        default=3,
    )
    old_formula_last = max(
        (
            row
            for row in range(4, sheet.max_row + 1)
            if any(sheet.cell(row, column).data_type == "f" for column in formulas)
        ),
        default=3,
    )
    target_last = 3 + len(records)
    clear_last = max(existing_last, old_formula_last, target_last)

    for row in range(4, clear_last + 1):
        for column in range(1, raw_width + 1):
            sheet.cell(row, column).value = None
        for column in formulas:
            sheet.cell(row, column).value = None

    for offset, record in enumerate(records, start=4):
        if offset > existing_last:
            copy_row_style(sheet, 4, offset, sheet.max_column)
        for column, value in enumerate(record, start=1):
            cell = sheet.cell(offset, column)
            cell.value = value
            header = sheet.cell(3, column).value
            if header in DATE_OUTPUT_HEADERS:
                cell.number_format = "dd/mm/yyyy"
            elif header in TIME_OUTPUT_HEADERS:
                cell.number_format = "hh:mm"
        for column, template_formula in formulas.items():
            destination = sheet.cell(offset, column)
            destination.value = Translator(
                template_formula,
                origin=sheet.cell(4, column).coordinate,
            ).translate_formula(destination.coordinate)

    if sheet.auto_filter.ref:
        from openpyxl.utils import get_column_letter

        sheet.auto_filter.ref = f"A3:{get_column_letter(sheet.max_column)}{max(target_last, 3)}"


def set_report_period(workbook, report_month: date) -> None:
    anchor = workbook["Checks"]["B20"]
    anchor.value = f'=TEXT(DATE({report_month.year},{report_month.month},1),"dd/mm/yyyy")'

    timeline = workbook["Performance Over Time"]
    for index, column in enumerate(range(2, 14)):
        month = report_month + relativedelta(months=index - 11)
        timeline.cell(5, column).value = datetime.combine(month, time())
        timeline.cell(5, column).number_format = "mmm yy"


def build_live_report(
    previous_live: bytes,
    new_rows: Sequence[Sequence[object]],
    mapping: "OrderedDict[str, Optional[str]]",
    store_db_records: Sequence[Dict[str, object]],
    report_month: date,
    expected_country: str,
) -> bytes:
    workbook = load_workbook_from_bytes(previous_live, data_only=False)
    validate_live_workbook(workbook, expected_country)
    raw_width = len(mapping)

    existing_ytd = raw_records(workbook, "YTD", raw_width)
    existing_r12m = raw_records(workbook, "R12M", raw_width)

    ytd_rows = [
        row
        for row in existing_ytd
        if record_visit_date(row) is not None and record_visit_date(row).year == report_month.year
    ]
    ytd_rows.extend(new_rows)

    r12m_start = report_month - relativedelta(months=11)
    r12m_end = report_month + relativedelta(months=1)
    r12m_rows = [
        row
        for row in existing_r12m
        if record_visit_date(row) is not None and r12m_start <= record_visit_date(row) < r12m_end
    ]
    r12m_rows.extend(new_rows)

    store_records = merged_store_records(
        workbook,
        store_db_records,
        [*ytd_rows, *r12m_rows],
        new_rows,
        mapping,
    )
    last_region_row = update_regions_sheet(workbook, store_records)
    update_region_lookup_ranges(workbook, last_region_row)
    make_abort_result_formulas_explicit(workbook)
    rebuild_store_performance(workbook, store_records)
    rebuild_district_performance(workbook, store_records)

    write_data_sheet(workbook["This Period"], new_rows, raw_width)
    write_data_sheet(workbook["YTD"], ytd_rows, raw_width)
    write_data_sheet(workbook["R12M"], r12m_rows, raw_width)
    set_report_period(workbook, report_month)

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    return workbook_to_bytes(
        workbook,
        drawing_source=previous_live,
        external_link_source=previous_live,
    )


def find_soffice() -> str:
    for candidate in ("libreoffice", "soffice"):
        path = shutil.which(candidate)
        if path:
            return path
    raise RuntimeError(
        "LibreOffice Calc is required to recalculate the reports. Add 'libreoffice-calc' to packages.txt."
    )


def recalculate_xlsx(data: bytes, filename: str) -> bytes:
    soffice = find_soffice()
    with tempfile.TemporaryDirectory(prefix="paddy_power_recalc_") as temp_dir:
        root = Path(temp_dir)
        source_dir = root / "source"
        output_dir = root / "output"
        profile_dir = root / "profile"
        source_dir.mkdir()
        output_dir.mkdir()
        profile_dir.mkdir()
        safe_name = Path(filename).name
        source_path = source_dir / safe_name
        source_path.write_bytes(data)
        profile_uri = profile_dir.resolve().as_uri()
        command = [
            soffice,
            "--headless",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            f"-env:UserInstallation={profile_uri}",
            "--convert-to",
            'xlsx:Calc MS Excel 2007 XML',
            "--outdir",
            str(output_dir),
            str(source_path),
        ]
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=240,
            check=False,
        )
        output_path = output_dir / safe_name
        if completed.returncode != 0 or not output_path.exists():
            detail = (completed.stderr or completed.stdout or "Unknown LibreOffice error").strip()
            raise RuntimeError(f"The report could not be recalculated: {detail}")
        return output_path.read_bytes()


def is_formula_cell(cell) -> bool:
    return cell.data_type == "f" or isinstance(cell.value, ArrayFormula) or (
        isinstance(cell.value, str) and cell.value.startswith("=")
    )


def build_non_live_report(formula_live: bytes, recalculated_live: bytes) -> bytes:
    formula_workbook = load_workbook_from_bytes(formula_live, data_only=False)
    value_workbook = load_workbook_from_bytes(recalculated_live, data_only=True)

    for sheet_name in VALUE_ONLY_SHEETS:
        if sheet_name not in formula_workbook.sheetnames:
            continue
        formula_sheet = formula_workbook[sheet_name]
        value_sheet = value_workbook[sheet_name]
        for cell in list(formula_sheet._cells.values()):
            if is_formula_cell(cell):
                cached_value = value_sheet[cell.coordinate].value
                if (
                    sheet_name == "Operational Questions"
                    and isinstance(cached_value, str)
                    and cached_value.startswith("#")
                ):
                    cached_value = None
                cell.value = cached_value

    summary = formula_workbook["Summary Data"]
    visit_count = value_workbook["Summary Data"]["C12"].value
    if isinstance(visit_count, (int, float)):
        desired_last_row = 16 + int(visit_count)
        if summary.max_row > desired_last_row:
            summary.delete_rows(desired_last_row + 1, summary.max_row - desired_last_row)

    for sheet in list(formula_workbook._sheets):
        if sheet.title not in NON_LIVE_SHEETS:
            formula_workbook.remove(sheet)

    # The retained report sheets have no external formulas, so exclude the
    # obsolete dashboard links entirely from the client-facing workbook.
    formula_workbook._external_links = []

    summary_index = formula_workbook.sheetnames.index("Summary Data")
    formula_workbook.active = summary_index
    if formula_workbook.views:
        formula_workbook.views[0].activeTab = summary_index
    for worksheet in formula_workbook.worksheets:
        worksheet.sheet_view.tabSelected = worksheet.title == "Summary Data"

    formula_workbook.calculation.calcMode = "auto"
    formula_workbook.calculation.fullCalcOnLoad = False
    formula_workbook.calculation.forceFullCalc = False
    return remove_invalid_defined_names(
        workbook_to_bytes(formula_workbook, drawing_source=formula_live)
    )


def zip_outputs(outputs: Dict[str, bytes]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for filename, data in outputs.items():
            archive.writestr(filename, data)
    return buffer.getvalue()


def validate_required_export_columns(df: pd.DataFrame) -> None:
    required = {"internal_id", "site_post_code", "date_of_visit"}
    missing = sorted(required - set(df.columns))
    if missing:
        raise ValueError(f"The data export is missing required column(s): {', '.join(missing)}.")


@st.cache_data(show_spinner=False)
def generate_reports(
    export_bytes: bytes,
    store_db_bytes: bytes,
    uk_live_bytes: bytes,
    roi_live_bytes: bytes,
    uk_live_name: str,
    roi_live_name: str,
) -> Tuple[Dict[str, bytes], Dict[str, int], str]:
    df = pd.read_csv(io.BytesIO(export_bytes), dtype=str, keep_default_na=False, encoding="utf-8-sig")
    validate_required_export_columns(df)
    store_database = load_store_database(store_db_bytes)
    original_export_count = len(df)
    df = df.drop_duplicates(subset=["internal_id"], keep="first").copy()
    duplicate_export_count = original_export_count - len(df)
    report_month = export_report_month(df)

    uk_template = load_workbook_from_bytes(uk_live_bytes, data_only=False)
    roi_template = load_workbook_from_bytes(roi_live_bytes, data_only=False)
    validate_live_workbook(uk_template, "UK")
    validate_live_workbook(roi_template, "ROI")

    uk_previous_month = report_month_from_workbook(uk_template, uk_live_name)
    roi_previous_month = report_month_from_workbook(roi_template, roi_live_name)
    expected_previous = report_month - relativedelta(months=1)
    if uk_previous_month and uk_previous_month != expected_previous:
        raise ValueError(
            f"The UK LIVE report is for {uk_previous_month.strftime('%B %Y')}; "
            f"the {report_month.strftime('%B %Y')} export requires the {expected_previous.strftime('%B %Y')} LIVE report."
        )
    if roi_previous_month and roi_previous_month != expected_previous:
        raise ValueError(
            f"The ROI LIVE report is for {roi_previous_month.strftime('%B %Y')}; "
            f"the {report_month.strftime('%B %Y')} export requires the {expected_previous.strftime('%B %Y')} LIVE report."
        )

    df["country_code"] = df["site_post_code"].map(classify_country)
    existing_uk_ids = audit_ids(uk_template, "YTD") | audit_ids(uk_template, "R12M")
    existing_roi_ids = audit_ids(roi_template, "YTD") | audit_ids(roi_template, "R12M")

    uk_df = df[
        (df["country_code"] == "GB")
        & (~df["internal_id"].astype(str).str.strip().isin(existing_uk_ids))
    ].copy()
    roi_df = df[
        (df["country_code"] == "IE")
        & (~df["internal_id"].astype(str).str.strip().isin(existing_roi_ids))
    ].copy()

    month_label = report_month.strftime("%B %Y")
    uk_live_name_out = f"Paddy Power UK Test Purchases - {month_label} - LIVE.xlsx"
    roi_live_name_out = f"Paddy Power ROI Test Purchases - {month_label} - LIVE.xlsx"
    uk_name_out = f"Paddy Power UK Test Purchases - {month_label}.xlsx"
    roi_name_out = f"Paddy Power ROI Test Purchases - {month_label}.xlsx"

    uk_unrecalculated = build_live_report(
        uk_live_bytes,
        mapped_rows(uk_df, GB_MAPPING),
        GB_MAPPING,
        store_database["UK"],
        report_month,
        "UK",
    )
    roi_unrecalculated = build_live_report(
        roi_live_bytes,
        mapped_rows(roi_df, IE_MAPPING),
        IE_MAPPING,
        store_database["ROI"],
        report_month,
        "ROI",
    )

    with ThreadPoolExecutor(max_workers=2) as executor:
        uk_future = executor.submit(recalculate_xlsx, uk_unrecalculated, uk_live_name_out)
        roi_future = executor.submit(recalculate_xlsx, roi_unrecalculated, roi_live_name_out)
        uk_recalculated = uk_future.result()
        roi_recalculated = roi_future.result()
    with ThreadPoolExecutor(max_workers=2) as executor:
        uk_non_live_future = executor.submit(
            build_non_live_report, uk_unrecalculated, uk_recalculated
        )
        roi_non_live_future = executor.submit(
            build_non_live_report, roi_unrecalculated, roi_recalculated
        )
        uk_non_live = uk_non_live_future.result()
        roi_non_live = roi_non_live_future.result()

    outputs = {
        uk_live_name_out: uk_unrecalculated,
        uk_name_out: uk_non_live,
        roi_live_name_out: roi_unrecalculated,
        roi_name_out: roi_non_live,
    }
    counts = {
        "UK": len(uk_df),
        "ROI": len(roi_df),
        "duplicates_omitted": duplicate_export_count + len(df) - len(uk_df) - len(roi_df),
    }
    return outputs, counts, month_label


def run_app() -> None:
    st.set_page_config(page_title="Paddy Power Report Generator", layout="centered")
    st.title("Paddy Power Report Generator")
    st.write(
        "Upload the new data export, the Paddy Power AV Store DB, and the previous month's UK and ROI LIVE reports. "
        "The generator will create updated LIVE and non-LIVE reports for both countries."
    )

    export_file = st.file_uploader("Upload data export", type="csv")
    store_db_file = st.file_uploader("Upload Paddy Power AV Store DB", type="xlsx")
    uk_live_file = st.file_uploader("Upload previous UK LIVE report", type="xlsx")
    roi_live_file = st.file_uploader("Upload previous ROI LIVE report", type="xlsx")

    if not (export_file and store_db_file and uk_live_file and roi_live_file):
        return

    try:
        with st.spinner("Generating and recalculating reports..."):
            outputs, counts, month_label = generate_reports(
                export_file.getvalue(),
                store_db_file.getvalue(),
                uk_live_file.getvalue(),
                roi_live_file.getvalue(),
                uk_live_file.name,
                roi_live_file.name,
            )
    except Exception as exc:
        st.error(str(exc))
        return

    st.success(
        f"{month_label} reports generated: {counts['UK']} UK visits and {counts['ROI']} ROI visits."
    )
    if counts["duplicates_omitted"]:
        st.info(f"{counts['duplicates_omitted']} previously reported or duplicate audit(s) were omitted.")

    st.download_button(
        "Download all four reports (.zip)",
        data=zip_outputs(outputs),
        file_name=f"Paddy Power Test Purchases - {month_label}.zip",
        mime="application/zip",
        type="primary",
    )

    st.subheader("Individual reports")
    for filename, data in outputs.items():
        st.download_button(
            f"Download {filename}",
            data=data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=filename,
        )


if __name__ == "__main__":
    run_app()
